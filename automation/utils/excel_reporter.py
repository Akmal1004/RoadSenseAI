import os
import json
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from automation.config.config import config

class ExcelReporter:
    COLORS = {
        'header': '1F3864',   # Dark blue
        'pass':   '70AD47',   # Green
        'fail':   'FF0000',   # Red
        'skip':   'FFC000',   # Amber
        'block':  '808080',   # Grey
        'title':  '2F75B6',   # Blue
        'white':  'FFFFFF',
        'light':  'D9E1F2',
    }

    def __init__(self, results: list, summary: dict):
        self.results = results
        self.summary = summary
        os.makedirs(os.path.join(config.REPORT_DIR, 'Excel'), exist_ok=True)

    def _header_style(self, cell, text, bold=True):
        cell.value = text
        cell.font = Font(bold=bold, color=self.COLORS['white'], size=11)
        cell.fill = PatternFill('solid', fgColor=self.COLORS['header'])
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    def _status_fill(self, status: str) -> PatternFill:
        color_map = {'PASSED': self.COLORS['pass'], 'FAILED': self.COLORS['fail'],
                     'SKIPPED': self.COLORS['skip'], 'BLOCKED': self.COLORS['block']}
        return PatternFill('solid', fgColor=color_map.get(status.upper(), self.COLORS['white']))

    def _set_column_widths(self, ws, widths: list):
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

    def generate_all(self) -> dict:
        paths = {}
        paths['main']    = self._generate_main_report()
        paths['passed']  = self._generate_filtered('Passed_Test_Cases.xlsx',  'PASSED')
        paths['failed']  = self._generate_filtered('Failed_Test_Cases.xlsx',  'FAILED')
        paths['summary'] = self._generate_summary_report()
        return paths

    def _generate_main_report(self) -> str:
        wb = Workbook()
        self._sheet_all_tests(wb.active)
        wb.active.title = 'All Test Cases'
        self._sheet_passed(wb.create_sheet('Passed Tests'))
        self._sheet_failed(wb.create_sheet('Failed Tests'))
        self._sheet_skipped(wb.create_sheet('Skipped Tests'))
        self._sheet_metrics(wb.create_sheet('Execution Metrics'))
        self._sheet_defects(wb.create_sheet('Defect Summary'))

        path = os.path.join(config.REPORT_DIR, 'Excel', 'Automation_Test_Report.xlsx')
        wb.save(path)
        return path

    def _sheet_all_tests(self, ws):
        headers = ['Test ID', 'Module', 'Test Name', 'Status', 'Duration(s)', 'Priority', 'Error Message', 'Screenshot']
        for i, h in enumerate(headers, 1):
            self._header_style(ws.cell(1, i), h)
        for r, t in enumerate(self.results, 2):
            ws.cell(r, 1, t.get('id', ''))
            ws.cell(r, 2, t.get('module', ''))
            ws.cell(r, 3, t.get('name', ''))
            c = ws.cell(r, 4, t.get('status', ''))
            c.fill = self._status_fill(t.get('status', ''))
            ws.cell(r, 5, round(t.get('duration', 0), 2))
            ws.cell(r, 6, t.get('priority', 'Medium'))
            ws.cell(r, 7, t.get('error', '')[:200] if t.get('error') else '')
            ws.cell(r, 8, t.get('screenshot', ''))
        self._set_column_widths(ws, [12, 18, 45, 10, 12, 10, 50, 40])

    def _sheet_passed(self, ws):
        passed = [t for t in self.results if t.get('status','').upper() == 'PASSED']
        headers = ['Test ID', 'Module', 'Test Name', 'Duration(s)', 'Priority']
        for i, h in enumerate(headers, 1): self._header_style(ws.cell(1,i), h)
        for r, t in enumerate(passed, 2):
            ws.cell(r,1,t.get('id','')); ws.cell(r,2,t.get('module',''))
            ws.cell(r,3,t.get('name','')); ws.cell(r,4,round(t.get('duration',0),2))
            ws.cell(r,5,t.get('priority','Medium'))
        self._set_column_widths(ws, [12,18,50,12,10])

    def _sheet_failed(self, ws):
        failed = [t for t in self.results if t.get('status','').upper() == 'FAILED']
        headers = ['Test ID','Module','Test Name','Error Message','Screenshot','Priority']
        for i, h in enumerate(headers, 1): self._header_style(ws.cell(1,i), h)
        for r, t in enumerate(failed, 2):
            ws.cell(r,1,t.get('id','')); ws.cell(r,2,t.get('module',''))
            ws.cell(r,3,t.get('name','')); ws.cell(r,4,t.get('error','')[:300] if t.get('error') else '')
            ws.cell(r,5,t.get('screenshot','')); ws.cell(r,6,t.get('priority','Medium'))
        self._set_column_widths(ws, [12,18,45,60,40,10])

    def _sheet_skipped(self, ws):
        skipped = [t for t in self.results if t.get('status','').upper() in ('SKIPPED','BLOCKED')]
        headers = ['Test ID','Module','Test Name','Reason']
        for i, h in enumerate(headers, 1): self._header_style(ws.cell(1,i), h)
        for r, t in enumerate(skipped, 2):
            ws.cell(r,1,t.get('id','')); ws.cell(r,2,t.get('module',''))
            ws.cell(r,3,t.get('name','')); ws.cell(r,4,t.get('skip_reason',''))
        self._set_column_widths(ws, [12,18,50,60])

    def _sheet_metrics(self, ws):
        ws.cell(1,1,'Metric'); ws.cell(1,2,'Value')
        for c in [ws.cell(1,1), ws.cell(1,2)]: self._header_style(c, c.value)
        metrics = [
            ('Total Test Cases', self.summary.get('total',0)),
            ('Executed', self.summary.get('executed',0)),
            ('Passed', self.summary.get('passed',0)),
            ('Failed', self.summary.get('failed',0)),
            ('Skipped', self.summary.get('skipped',0)),
            ('Pass Rate', f"{self.summary.get('pass_rate',0):.1f}%"),
            ('Total Duration (s)', f"{self.summary.get('duration',0):.2f}"),
            ('Execution Date', self.summary.get('date', '')),
            ('Environment', config.BASE_URL),
        ]
        for r, (k, v) in enumerate(metrics, 2):
            ws.cell(r,1,k); ws.cell(r,2,str(v))
        self._set_column_widths(ws, [30,40])

    def _sheet_defects(self, ws):
        failed = [t for t in self.results if t.get('status','').upper() == 'FAILED']
        headers = ['Defect ID','Test ID','Module','Defect Description','Priority','Status']
        for i, h in enumerate(headers, 1): self._header_style(ws.cell(1,i), h)
        for r, t in enumerate(failed, 2):
            ws.cell(r,1,f"DEF-{r-1:03d}"); ws.cell(r,2,t.get('id',''))
            ws.cell(r,3,t.get('module','')); ws.cell(r,4,t.get('error','')[:200] if t.get('error') else 'See logs')
            ws.cell(r,5,t.get('priority','Medium')); ws.cell(r,6,'Open')
        self._set_column_widths(ws, [12,12,18,60,10,12])

    def _generate_filtered(self, filename: str, status: str) -> str:
        filtered = [t for t in self.results if t.get('status','').upper() == status.upper()]
        wb = Workbook()
        ws = wb.active
        ws.title = f'{status.capitalize()} Tests'
        headers = ['Test ID','Module','Test Name','Duration(s)','Priority','Error']
        for i, h in enumerate(headers, 1): self._header_style(ws.cell(1,i), h)
        for r, t in enumerate(filtered, 2):
            ws.cell(r,1,t.get('id','')); ws.cell(r,2,t.get('module',''))
            ws.cell(r,3,t.get('name','')); ws.cell(r,4,round(t.get('duration',0),2))
            ws.cell(r,5,t.get('priority','Medium')); ws.cell(r,6,t.get('error','')[:200] if t.get('error') else '')
        self._set_column_widths(ws, [12,18,50,12,10,60])
        path = os.path.join(config.REPORT_DIR, 'Excel', filename)
        wb.save(path)
        return path

    def _generate_summary_report(self) -> str:
        wb = Workbook()
        ws = wb.active
        ws.title = 'Summary'
        ws.merge_cells('A1:F1')
        ws['A1'] = 'RoadSenseAI – Automation Test Execution Summary'
        ws['A1'].font = Font(bold=True, size=14, color=self.COLORS['white'])
        ws['A1'].fill = PatternFill('solid', fgColor=self.COLORS['title'])
        ws['A1'].alignment = Alignment(horizontal='center')
        
        data = [
            ('Metric', 'Value'),
            ('Total Tests', self.summary.get('total', 0)),
            ('Passed', self.summary.get('passed', 0)),
            ('Failed', self.summary.get('failed', 0)),
            ('Skipped', self.summary.get('skipped', 0)),
            ('Pass Rate', f"{self.summary.get('pass_rate', 0):.1f}%"),
            ('Duration (s)', f"{self.summary.get('duration', 0):.2f}"),
            ('Date', self.summary.get('date', '')),
            ('Base URL', config.BASE_URL),
        ]
        for r, (k, v) in enumerate(data, 2):
            ws.cell(r,1,k); ws.cell(r,2,str(v))
            if r == 2:
                self._header_style(ws.cell(r,1), k)
                self._header_style(ws.cell(r,2), str(v))
        self._set_column_widths(ws, [30, 40])
        path = os.path.join(config.REPORT_DIR, 'Excel', 'Summary_Report.xlsx')
        wb.save(path)
        return path
